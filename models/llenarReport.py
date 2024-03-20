from openpyxl import Workbook
from models.database import Database
from sqlalchemy import text
from openpyxl.styles import Alignment,  Font

class LlenarReporte:
    def __init__(self, ruta_reporte):
        self.ruta_reporte = ruta_reporte
        self.libro = Workbook()
        self.hoja_usuarios = self.libro.active
        self.hoja_usuarios.title = "USUARIOS"
        self.hoja_zonas = self.libro.create_sheet(title="ZONAS")
        self.hoja_cabanas = self.libro.create_sheet(title="CABANAS")



    def llenar_celda(self, hoja, celda, valor):
        hoja[celda] = valor

    def llenar_hoja_usuarios(self, registros):
        
        self.hoja_usuarios.merge_cells('A1:G7')
        
        titulo_celda = self.hoja_usuarios['A1']
        titulo_celda.value = 'RESUMEN USUARIOS'
       
        titulo_celda.font = Font(size=14, bold=True)
        
        titulo_celda.alignment = Alignment(horizontal='center', vertical='center')

        
        # Encabezados
        encabezados = ["ID", "Nombre", "Apellido Paterno", "Apellido Materno", "Alias", "Email", "Tipo Usuario"]
        for col_num, encabezado in enumerate(encabezados, 1):
            self.hoja_usuarios.cell(row=8, column=col_num, value=encabezado)

        # Datos
        for row_num, registro in enumerate(registros, 9):  # Comienza desde la fila 9
            for col_num, valor in enumerate(registro, 1):
                self.hoja_usuarios.cell(row=row_num, column=col_num, value=valor)

    def llenar_hoja_zonas(self, registros):
        self.hoja_zonas.merge_cells('A1:G7')
        
        titulo_celda = self.hoja_zonas['A1']
        titulo_celda.value = 'RESUMEN ZONAS'
       
        titulo_celda.font = Font(size=14, bold=True)
        
        titulo_celda.alignment = Alignment(horizontal='center', vertical='center')
        # Encabezados
        encabezados = ["ID", "Nombre", "Ubicación", "Activo", "ID Usuario", "Fecha de Actualización"]
        for col_num, encabezado in enumerate(encabezados, 1):
            self.hoja_zonas.cell(row=8, column=col_num, value=encabezado)

        # Datos
        for row_num, registro in enumerate(registros, 9):  # Comienza desde la fila 9
            for col_num, valor in enumerate(registro, 1):
                self.hoja_zonas.cell(row=row_num, column=col_num, value=valor)

    def llenar_hoja_cabanas_por_zonas(self, registros):
        self.hoja_cabanas.merge_cells('A1:G7')
        
        titulo_celda = self.hoja_cabanas['A1']
        titulo_celda.value = 'RESUMEN CABAÑAS'
       
        titulo_celda.font = Font(size=14, bold=True)
        
        titulo_celda.alignment = Alignment(horizontal='center', vertical='center')
        # Encabezados
        encabezados = ["Nombre Zona", "Cabañas en Zona"]
        for col_num, encabezado in enumerate(encabezados, 1):
            self.hoja_cabanas.cell(row=8, column=col_num, value=encabezado)

        # Datos
        for row_num, registro in enumerate(registros, 9):
            nombre_zona, cabañas_en_zona = registro
            self.hoja_cabanas.cell(row=row_num, column=1, value=nombre_zona)

            # Manejar el caso en que cabañas_en_zona es None o vacío
            cabañas_en_zona = [str(cabana) for cabana in cabañas_en_zona if cabana is not None]
            
            if cabañas_en_zona:
                self.hoja_cabanas.cell(row=row_num, column=2, value=", ".join(cabañas_en_zona))
            else:
                self.hoja_cabanas.cell(row=row_num, column=2, value="No tiene registros")

    def llenar_hoja_resumen(self, usuarios_total, zonas_total, cabanas_total, fechas_dis_total, fechas_nodis_total):
        hoja_resumen = self.libro.create_sheet(title="RESUMEN")
        hoja_resumen.merge_cells('A1:G7')
        
        titulo_celda = hoja_resumen['A1']
        titulo_celda.value = 'RESUMEN GENERAL'
       
        titulo_celda.font = Font(size=14, bold=True)
        
        titulo_celda.alignment = Alignment(horizontal='center', vertical='center')
        
        

        # Títulos y resultados
        titulos_resultados = [
            ("TOTAL DE USUARIOS", usuarios_total),
            ("TOTAL DE ZONAS", zonas_total),
            ("TOTAL DE CABAÑAS", cabanas_total),
            ("TOTAL DE FECHAS DISPONIBLES", fechas_dis_total),
            ("TOTAL DE FECHAS OCUPADAS", fechas_nodis_total),
        ]

        # Iterar sobre los títulos y resultados
        for i, (titulo, resultado) in enumerate(titulos_resultados, 1):
            hoja_resumen.cell(row=i + 7, column=1, value=titulo).alignment = Alignment(horizontal='right')
            hoja_resumen.cell(row=i + 7, column=2, value=resultado)


    
    def guardar_reporte(self):
        self.libro.save(self.ruta_reporte)

    def tabla_usuarios():
            try:
                db = Database()
                conn = db.engine.connect()

                # Consulta a la tabla de usuarios
                query = text("""
                    SELECT u.id_usr, 
                        u.nombre,
                        u.apellidoP,
                        u.apellidoM,
                        u.alias,
                        u.email,
                        t.tipo_usr
                    FROM usuarios u
                    JOIN tipo_usuario t ON u.id_tuser = t.id_tpurs
                """)

                result = conn.execute(query)
                usuarios = result.fetchall()

                conn.close()

                return usuarios

            except Exception as e:
                print(f"Error al obtener la tabla de usuarios: {e}")
                return []

    def tabla_zonas():
            try:
                db = Database()
                conn = db.engine.connect()

                # Consulta a la tabla de zonas con JOIN en la tabla de usuarios
                query = text("""
                    SELECT z.id_zn,
                        z.nombre_zn,
                        z.ubicacion_zn,
                        z.activo_zn,
                        u.nombre || ' ' || u.apellidoP || ' ' || u.apellidoM AS nombre_usuario,
                        z.uptade_zn
                    FROM zonas z
                    JOIN usuarios u ON z.id_usr = u.id_usr
                """)

                result = conn.execute(query)
                zonas = result.fetchall()

                conn.close()

                return zonas

            except Exception as e:
                print(f"Error al obtener la tabla de zonas: {e}")
                return []
    
    def obtener_cabanas_por_zonas():
        try:
            db = Database()
            conn = db.engine.connect()

            # Consulta para obtener cabañas agrupadas por zonas
            query = text("""
                SELECT z.nombre_zn AS nombre_zona,
                    ARRAY_AGG(c.no_cbn) AS cabañas_en_zona
                FROM zonas z
                LEFT JOIN cabanas c ON z.id_zn = c.id_zn
                GROUP BY z.nombre_zn
            """)

            result = conn.execute(query)
            cabanas_por_zonas = result.fetchall()

            conn.close()

            return cabanas_por_zonas

        except Exception as e:
            print(f"Error al obtener las cabañas por zonas: {e}")
            return []
    
    def contar_registros_usuarios():
        try:
            db = Database()
            conn = db.engine.connect()

            # Consulta para contar los registros en la tabla de usuarios
            query = text("""
                SELECT COUNT(*) FROM usuarios
            """)

            result = conn.execute(query)
            count = result.scalar()  # Obtenemos el resultado escalar

            conn.close()

            return count

        except Exception as e:
            print(f"Error al contar los registros de usuarios: {e}")
            return 0  # En caso de error, retornamos 0   
    
    def contar_registros_zonas():
        try:
            db = Database()
            conn = db.engine.connect()

            # Consulta para contar los registros en la tabla de zonas
            query = text("""
                SELECT COUNT(*) FROM zonas
            """)

            result = conn.execute(query)
            count = result.scalar()  # Obtenemos el resultado escalar

            conn.close()

            return count

        except Exception as e:
            print(f"Error al contar los registros de zonas: {e}")
            return 0  # En caso de error, retornamos 0
    
    def contar_registros_cabanas():
        try:
            db = Database()
            conn = db.engine.connect()

            # Consulta para contar los registros en la tabla de cabanas
            query = text("""
                SELECT COUNT(*) FROM cabanas
            """)

            result = conn.execute(query)
            count = result.scalar()  # Obtenemos el resultado escalar

            conn.close()

            return count

        except Exception as e:
            print(f"Error al contar los registros de cabanas: {e}")
            return 0  # En caso de error, retornamos 0
    
    def contar_fechas_disponibles():
        try:
            db = Database()
            conn = db.engine.connect()

            # Consulta para contar las fechas disponibles (estado 1)
            query = text("""
                SELECT COUNT(*) FROM fechas
                WHERE disponible = B'1'
            """)

            result = conn.execute(query)
            count = result.scalar()  # Obtenemos el resultado escalar

            conn.close()

            return count

        except Exception as e:
            print(f"Error al contar las fechas disponibles: {e}")
            return 0  # En caso de error, retornamos 0

    def contar_fechas_no_disponibles():
        try:
            db = Database()
            conn = db.engine.connect()

            # Consulta para contar las fechas no disponibles (estado 0)
            query = text("""
                SELECT COUNT(*) FROM fechas
                WHERE disponible = B'0'
            """)

            result = conn.execute(query)
            count = result.scalar()  # Obtenemos el resultado escalar

            conn.close()

            return count

        except Exception as e:
            print(f"Error al contar las fechas no disponibles: {e}")
            return 0  # En caso de error, retornamos 0

    #